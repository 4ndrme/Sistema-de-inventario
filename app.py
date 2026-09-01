# --- LIBRERÍAS ESTÁNDAR ---
import os
import io
import smtplib
from email.mime.text import MIMEText

# --- LIBRERÍAS DE TERCEROS ---
from flask import Flask, render_template, request, redirect, url_for, flash, session, Response
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- MÓDULOS LOCALES ---
from config import Config
from models import db, Usuario, Producto, ProductoFisico, ProductoPerecible, Movimiento, ConfiguracionSistema, ProductoDigital
from permisos import requiere_rol

# --- INICIALIZACIÓN DE LA APP ---
app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

# --- GUARDIÁN DE SEGURIDAD (Protección de Rutas) ---
@app.before_request
def bloquear_accesos_no_autorizados():
    # Permitimos 'login', 'registro' y archivos estáticos sin iniciar sesión
    rutas_libres = ['login', 'registro', 'static']
    
    # Si el usuario no ha iniciado sesión y la ruta es privada, redirigimos a login
    if 'usuario_id' not in session and request.endpoint not in rutas_libres:
        return redirect(url_for('login'))

# --- RUTA DE REGISTRO DE USUARIOS ---
@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        username = request.form.get("username").strip() 
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")
        rol_seleccionado = request.form.get("rol") 
        
        # 1. Validaciones de longitud
        if len(username) < 4:
            flash("El nombre de usuario debe tener al menos 4 caracteres.", "error")
            return redirect(url_for("registro"))
            
        if len(password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres por seguridad.", "error")
            return redirect(url_for("registro"))

        # 2. Validación de coincidencia
        if password != confirm_password:
            flash("Las contraseñas no coinciden.", "error")
            return redirect(url_for("registro"))
            
        # 3. Validación de existencia
        usuario_existente = Usuario.query.filter_by(username=username).first()
        if usuario_existente:
            flash("El nombre de usuario ya está registrado.", "error")
            return redirect(url_for("registro"))
            
        nuevo_usuario = Usuario(username=username, rol=rol_seleccionado)
        nuevo_usuario.set_password(password)
        
        db.session.add(nuevo_usuario)
        db.session.commit()

        flash("Registro exitoso. Ahora puedes iniciar sesión.", "success")
        return redirect(url_for("login"))

    return render_template("registro.html")

# --- RUTA DE AUTENTICACIÓN (LOGIN) ---
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        
        usuario = Usuario.query.filter_by(username=username).first()
        
        # Comparamos el hash seguro
        if usuario and usuario.check_password(password):
            session["usuario_id"] = usuario.id
            session["username"] = usuario.username
            session["rol"] = usuario.rol
            flash(f"Bienvenido al sistema, {usuario.username}.", "success")
            return redirect(url_for("inicio"))
        else:
            flash("Usuario o contraseña incorrectos.", "error")
            return redirect(url_for("login")) 
            
    return render_template("login.html")

# --- RUTA DE CIERRE DE SESIÓN ---
@app.route("/logout")
def logout():
    session.clear()
    flash("Has cerrado sesión correctamente.", "success")
    return redirect(url_for("login"))

# --- RUTA PRINCIPAL (DASHBOARD) ---
@app.route("/")
def inicio():
    query = request.args.get("q", "")
    if query:
        productos = Producto.query.filter(
            (Producto.nombre.ilike(f"%{query}%")) | 
            (Producto.codigo.ilike(f"%{query}%"))
        ).all()
    else:
        productos = Producto.query.all()
        
    config = ConfiguracionSistema.query.first() 
    
    # --- NUEVA DEFENSA: Autogenerar si la base de datos está vacía ---
    if not config:
        config = ConfiguracionSistema()
        db.session.add(config)
        db.session.commit()

    return render_template("index.html", productos=productos, query=query, config=config)

# --- RUTA DE CREACIÓN DE MATERIALES ---
@app.route("/nuevo_material", methods=["GET", "POST"])
def nuevo_material():
    if request.method == "POST":
        codigo = request.form.get("codigo")
        nombre = request.form.get("nombre")
        tipo = request.form.get("tipo")
        stock = request.form.get("stock")
        stock_int = int(stock) if stock else 0

        # POLIMORFISMO EN ACCIÓN: Instanciamos la clase hija correspondiente
        if tipo == 'Perecible':
            fecha_caducidad = request.form.get("fecha_caducidad")
            
            # Candado de seguridad backend
            if not fecha_caducidad:
                flash("Error de integridad: Los materiales perecibles exigen obligatoriamente una fecha de caducidad.", "error")
                return redirect(url_for('nuevo_material'))
            
            nuevo_prod = ProductoPerecible(
                codigo=codigo,
                nombre=nombre,
                tipo=tipo,
                stock=stock_int,
                fecha_caducidad=fecha_caducidad
            )
            
        elif tipo == 'Digital':
            # --- LOGICA PARA DIGITALES ---
            nuevo_prod = ProductoDigital(
                codigo=codigo,
                nombre=nombre,
                tipo=tipo,
                stock=stock_int,
                enlace_descarga=None #Escalabilidad
            )
            
        else:
            # Por defecto, Producto Físico
            nuevo_prod = ProductoFisico(
                codigo=codigo,
                nombre=nombre,
                tipo=tipo,
                stock=stock_int,
                ruta_documento=None #Escalabilidad
            )
        
        db.session.add(nuevo_prod)
        try:
            db.session.commit()
            flash(f"Material {nombre} registrado correctamente.", "success")
            return redirect(url_for('inicio'))
            
        except IntegrityError:
            db.session.rollback() 
            flash("Error: El código ingresado ya existe en la base de datos. Intente con uno distinto.", "error")
            return redirect(url_for('nuevo_material'))
        
    return render_template("nuevo_material.html")

# --- RUTA DE ELIMINACIÓN DE PRODUCTOS ---
@app.route("/eliminar/<int:id>", methods=["POST"])
def eliminar_producto(id):
    producto = db.session.get(Producto, id)
    if producto:
        try:
            db.session.delete(producto)
            db.session.commit()
            flash("Material eliminado correctamente.", "success")
        except IntegrityError:
            db.session.rollback() # Deshace el intento de borrado para evitar que la BD se trabe
            flash("Acción bloqueada: No se puede eliminar un material que posee historial de movimientos por motivos de auditoría.", "error")
    return redirect(url_for("inicio"))

# --- RUTA DE TRANSACCIONES DE INVENTARIO ---
@app.route("/procesar_movimiento/<int:id_producto>", methods=["GET", "POST"])
def procesar_movimiento(id_producto):
    producto = Producto.query.get_or_404(id_producto)
    
    if request.method == "POST":
        tipo_movimiento = request.form.get("tipo_movimiento") 
        cantidad = int(request.form.get("cantidad", 0))
        observacion = request.form.get("observacion", "")
        usuario_id = session["usuario_id"]

        try:
            # Delegamos la transacción al Stored Procedure en SQL Server
            sql = text("""
                EXEC sp_ProcesarMovimiento 
                    @producto_id = :prod_id, 
                    @usuario_id = :usr_id, 
                    @tipo_movimiento = :tipo, 
                    @cantidad = :cant, 
                    @observacion = :obs
            """)
            
            db.session.execute(sql, {
                'prod_id': producto.id,
                'usr_id': usuario_id,
                'tipo': tipo_movimiento,
                'cant': cantidad,
                'obs': observacion 
            })
            db.session.commit()
            
            # Refrescamos el objeto en memoria para leer el nuevo stock
            db.session.refresh(producto)

            # Utilizamos .upper() para blindarnos contra diferencias de mayúsculas/minúsculas desde el HTML
            if tipo_movimiento.upper() == "DESPACHO" and producto.requiere_atencion():
                enviar_alerta_stock(producto) 
                # Notificación visual específica para alertar al operador en pantalla
                flash(f"Despacho procesado. ALERTA ENVIADA: El stock actual ({producto.stock}) cayó por debajo del límite.", "error")
            else:
                flash(f"{tipo_movimiento} de {cantidad} unidades procesado con éxito.", "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error de base de datos: La operación fue bloqueada por regla de negocio.", "error")
            
        return redirect(url_for('inicio'))
        
    return render_template("movimiento.html", producto=producto)

# --- RUTA DE HISTORIAL DE MOVIMIENTOS ---
@app.route("/movimientos")
def movimientos():
    movimientos_lista = Movimiento.query.order_by(Movimiento.fecha_movimiento.desc()).all()
    return render_template("movimientos.html", movimientos=movimientos_lista)

# --- RUTA DEL PANEL DE CONTROL GLOBAL ---
@app.route("/configuracion", methods=["GET", "POST"])
@requiere_rol("Supervisor")
def configuracion():
    config = ConfiguracionSistema.query.first()
    if not config:
        config = ConfiguracionSistema()
        db.session.add(config)
        db.session.commit()

    if request.method == "POST":
        config.correo_alertas = request.form.get("correo_alertas")
        config.dias_alerta_caducidad = int(request.form.get("dias_alerta_caducidad", 30))
        config.umbral_stock_critico = int(request.form.get("umbral_stock_critico", 50))
        
        db.session.commit()
        flash("Parámetros globales actualizados correctamente.", "success")
        return redirect(url_for('configuracion'))
        
    return render_template("configuracion.html", config=config)

# --- RUTA DE EXPORTACIÓN A EXCEL ---
@app.route("/exportar_inventario")
def exportar_inventario():
    productos = Producto.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario WMS"

    headers = ['Código del Material', 'Nombre', 'Tipo', 'Stock Actual', 'Caducidad', 'Estado']
    ws.append(headers)

    fill_teal = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    font_bold_white = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = fill_teal
        cell.font = font_bold_white
        cell.alignment = align_center

    for p in productos:
        fecha_segura = getattr(p, 'fecha_caducidad', None)
        caducidad = fecha_segura.strftime('%Y-%m-%d') if fecha_segura else 'N/A'
        estado = 'Activo' if p.activo else 'Inactivo'
        ws.append([p.codigo, p.nombre, p.tipo, p.stock, caducidad, estado])

    column_widths = {'A': 20, 'B': 45, 'C': 15, 'D': 15, 'E': 15, 'F': 12}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = Response(output.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename=Reporte_Inventario_WMS.xlsx'
    
    return response

# --- MOTOR DE ALERTAS SMTP ---
def enviar_alerta_stock(producto):
    remitente = os.getenv("EMAIL_USER") 
    password = os.getenv("EMAIL_PASS") 
    
    if not remitente or not password:
        print("Error: Credenciales SMTP no configuradas en las variables de entorno.")
        return

    config = ConfiguracionSistema.query.first()
    destinatario = config.correo_alertas if config else "michaelandresqc@gmail.com"

    detalle_alerta = producto.generar_mensaje_alerta()

    mensaje_cuerpo = (
        f"ALERTA WMS AUTOMÁTICA\n\n"
        f"{detalle_alerta}\n\n"
        f"Por favor, proceda con las acciones de reabastecimiento o revisión correspondientes."
    )
    
    msg = MIMEText(mensaje_cuerpo)
    msg['Subject'] = f"⚠️ Notificación WMS: {producto.nombre}"
    msg['From'] = remitente
    msg['To'] = destinatario

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(remitente, password)
        server.send_message(msg)
        server.quit()
        print(f"Alerta enviada exitosamente para {producto.nombre}")
    except Exception as e:
        print(f"Error al enviar la alerta SMTP: {e}")

# --- RUTA DE INTELIGENCIA DE NEGOCIOS (REPORTES) ---
@app.route("/reportes")
@requiere_rol("Auditor", "Supervisor")
def reportes():
    sql = text("""
        SELECT
            p.codigo AS Codigo_Material,
            p.nombre AS Material,
            p.tipo AS Categoria,
            p.stock AS Stock_Actual,
            COUNT(m.id) AS Total_Transacciones,
            ISNULL(SUM(CASE WHEN m.tipo_movimiento = 'INGRESO' THEN m.cantidad ELSE 0 END), 0) AS Total_Ingresado,
            ISNULL(SUM(CASE WHEN m.tipo_movimiento = 'DESPACHO' THEN m.cantidad ELSE 0 END), 0) AS Total_Despachado
        FROM productos p
        LEFT JOIN movimientos m ON p.id = m.producto_id
        GROUP BY p.codigo, p.nombre, p.tipo, p.stock
        ORDER BY Total_Transacciones DESC;
    """)
    
    resultados = db.session.execute(sql).fetchall()
    return render_template("reportes.html", estadisticas=resultados)

if __name__ == "__main__":
    app.run(debug=True)